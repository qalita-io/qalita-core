"""
# QALITA (c) COPYRIGHT 2025 - ALL RIGHTS RESERVED -

Loader-side streaming helpers.

``data_source_opener`` turns a source into Parquet parts; this module holds the
mechanics it uses to do that without ever holding a dataset in memory. Three
invariants matter more than anything else here:

1. **One pinned Arrow schema per logical object.** A writer that infers its own
   dtypes per part lets ``*_part_1`` type a column ``Int64`` and ``*_part_7``
   ``String``; ``pl.scan_parquet(parts)`` then raises ``SchemaError`` on exactly
   the large datasets this module exists for. The schema is decided once, up
   front, and every batch is cast to it.

2. **Nothing accumulates.** Every writer here consumes an iterator and forgets
   each batch as soon as it is on disk. The only bounded buffer is one chunk.

3. **Disk is checked before it is filled.** A 100 GiB source stages to tens of
   GiB; running the volume to zero mid-load leaves a half-written dataset and an
   unusable worker.

The previous version of this module exported fake-lazy helpers (``scan_excel``,
``scan_json`` and ``read_database_streaming`` all read the whole source, then
called ``.lazy()`` on it) plus ``stream_to_parquet_chunks``, which re-executed
the query plan once per slice — writing N parts meant N full scans of the
source. Nothing imported any of them, and ``data_source_opener`` had grown its
own drifting copies instead; those copies now live here, once.
"""

from __future__ import annotations

import datetime as _dt
import json
import logging
import os
import shutil
from decimal import Decimal
from pathlib import Path
from typing import Any, Dict, Iterable, Iterator, List, Optional, Sequence

import polars as pl
import pyarrow as pa
import pyarrow.parquet as pq

logger = logging.getLogger(__name__)

__all__ = [
    "DEFAULT_CHUNK_ROWS",
    "DEFAULT_FETCH_ROWS",
    "DEFAULT_MIN_FREE_BYTES",
    "InsufficientDiskSpaceError",
    "ParquetPartWriter",
    "SchemaDriftError",
    "arrow_type_hints",
    "check_disk_space",
    "iter_excel_rows",
    "iter_json_array",
    "part_path",
    "scan_csv",
    "scan_ndjson",
    "scan_parquet",
    "sink_parts",
    "sniff_json_format",
    "stream_sql_to_parquet",
    "stream_to_parquet",
    "write_arrow_batches",
    "write_dict_rows",
    "write_row_batches",
]


# Rows per Parquet part file, and per row group inside it. Parts are what makes
# a load restartable and what bounds the writer's memory.
DEFAULT_CHUNK_ROWS = 100_000

# Rows pulled from a driver at a time. Smaller than a part on purpose: this is
# what is alive in Python, and a driver row costs an order of magnitude more
# there than in an Arrow buffer.
DEFAULT_FETCH_ROWS = 10_000

# Refuse to start (or to open another part) with less than this much free space
# on the staging volume. Two thirds of a GiB is roughly one 100k-row part of a
# wide table plus the room Parquet needs to close it.
DEFAULT_MIN_FREE_BYTES = 512 * 1024 * 1024

# Compression is pinned so that every path in the loader produces the same kind
# of file. The pandas path used to write snappy silently while the Polars path
# wrote zstd, which made part files of the same object differ in size by 3x.
PARQUET_COMPRESSION = "zstd"


class InsufficientDiskSpaceError(RuntimeError):
    """The staging volume cannot hold what is about to be written."""


class SchemaDriftError(RuntimeError):
    """A batch could not be cast to the object's pinned schema.

    Raised instead of writing a part with a different schema: the divergence
    would only surface later, as a ``SchemaError`` from ``pl.scan_parquet`` over
    the parts, with nothing left to say which part introduced it.
    """


# ---------------------------------------------------------------------------
# Paths and disk
# ---------------------------------------------------------------------------


def part_path(output_dir: str, base_name: str, part_index: int) -> str:
    """Path of one part file. ``Pack`` recovers the object name from this."""
    return os.path.join(output_dir, f"{base_name}_part_{part_index}.parquet")


def check_disk_space(
    directory: str,
    estimated_bytes: Optional[int] = None,
    *,
    min_free_bytes: int = DEFAULT_MIN_FREE_BYTES,
) -> int:
    """Refuse to stage data that will not fit, before writing a single byte.

    Returns the number of free bytes so callers can log it.
    """
    Path(directory).mkdir(parents=True, exist_ok=True)
    free = shutil.disk_usage(directory).free

    if estimated_bytes and free < estimated_bytes:
        raise InsufficientDiskSpaceError(
            f"staging to {directory} needs about "
            f"{estimated_bytes / 2**30:.1f} GiB but only "
            f"{free / 2**30:.1f} GiB are free. Point "
            f"'parquet_output_dir' at a larger volume, or restrict the load "
            f"with a table name or a SQL query."
        )
    if free < min_free_bytes:
        raise InsufficientDiskSpaceError(
            f"only {free / 2**20:.0f} MiB free on the volume holding "
            f"{directory}; at least {min_free_bytes / 2**20:.0f} MiB are "
            f"required to stage a source safely."
        )
    return free


# ---------------------------------------------------------------------------
# Schema pinning
# ---------------------------------------------------------------------------


# Python types are what every DB-API driver and every SQLAlchemy dialect agree
# on, so the hint is derived from those rather than from dialect-specific SQL
# types. The Arrow types chosen here are the ones Polars itself produces, which
# keeps the common case cast-free.
_PYTHON_TO_ARROW = {
    bool: pa.bool_(),
    int: pa.int64(),
    float: pa.float64(),
    str: pa.large_string(),
    bytes: pa.large_binary(),
    _dt.datetime: pa.timestamp("us"),
    _dt.date: pa.date32(),
    _dt.time: pa.time64("us"),
}


def arrow_type_hints(columns: Iterable[Any]) -> Dict[str, "pa.DataType"]:
    """Arrow types for SQLAlchemy column descriptions, best effort.

    Used to type columns that are entirely NULL in the first batch. Without it
    such a column is inferred as Arrow ``null`` and every later batch that
    actually has values fails to cast into it.
    """
    hints: Dict[str, pa.DataType] = {}
    for column in columns or []:
        try:
            name = column["name"]
            python_type = column["type"].python_type
        except Exception:  # noqa: BLE001 - exotic dialect types have none
            continue
        arrow_type = _PYTHON_TO_ARROW.get(python_type)
        if arrow_type is not None:
            hints[name] = arrow_type
    return hints


def _pin_schema(
    schema: "pa.Schema",
    type_hints: Optional[Dict[str, "pa.DataType"]] = None,
) -> "pa.Schema":
    """Resolve the schema every part of an object will be written with.

    Only columns inferred as ``null`` are overridden: an all-null first batch
    says nothing about the column, whereas an inferred concrete type is evidence
    from the data itself and outranks any metadata hint.
    """
    hints = type_hints or {}
    fields = []
    for field in schema:
        if pa.types.is_null(field.type):
            fields.append(
                pa.field(field.name, hints.get(field.name, pa.large_string()))
            )
        else:
            fields.append(field)
    return pa.schema(fields)


def _conform(table: "pa.Table", schema: "pa.Schema") -> "pa.Table":
    """Cast a batch to the pinned schema, or say precisely what diverged."""
    if table.schema.equals(schema):
        return table

    missing = [n for n in schema.names if n not in table.schema.names]
    if missing:
        raise SchemaDriftError(
            f"batch is missing column(s) {missing} present in the pinned "
            f"schema {schema.names}"
        )

    try:
        return table.select(schema.names).cast(schema)
    except Exception as exc:  # noqa: BLE001 - re-raised with the diff below
        diff = [
            f"{f.name}: {table.schema.field(f.name).type} -> {f.type}"
            for f in schema
            if table.schema.field(f.name).type != f.type
        ]
        raise SchemaDriftError(
            f"a batch could not be cast to the schema pinned for this object "
            f"({'; '.join(diff) or exc}). Load the object with an explicit "
            f"query that casts the column, or split it into two objects."
        ) from exc


def _batch_rows(chunk_rows: int, fetch_rows: Optional[int]) -> int:
    """Rows held in Python at once, never more than one part."""
    return max(1, min(int(chunk_rows), int(fetch_rows or DEFAULT_FETCH_ROWS)))


def _as_table(data: Any) -> "pa.Table":
    if isinstance(data, pa.Table):
        return data
    if isinstance(data, pa.RecordBatch):
        return pa.Table.from_batches([data])
    if isinstance(data, pl.DataFrame):
        return data.to_arrow()
    raise TypeError(
        f"expected a pyarrow Table/RecordBatch or a polars DataFrame, got "
        f"{type(data).__name__}"
    )


# ---------------------------------------------------------------------------
# The writer
# ---------------------------------------------------------------------------


class ParquetPartWriter:
    """Append Arrow batches to ``<base>_part_<n>.parquet`` files.

    One ``ParquetWriter`` is open at a time and it holds a single schema, so the
    parts of an object are readable as one dataset by construction. Parts roll
    at exactly ``chunk_rows`` rows, which keeps the file names meaningful and
    the peak memory equal to one batch.
    """

    def __init__(
        self,
        output_dir: str,
        base_name: str,
        *,
        schema: Optional["pa.Schema"] = None,
        type_hints: Optional[Dict[str, "pa.DataType"]] = None,
        chunk_rows: int = DEFAULT_CHUNK_ROWS,
        compression: str = PARQUET_COMPRESSION,
        min_free_bytes: int = DEFAULT_MIN_FREE_BYTES,
    ) -> None:
        self.output_dir = output_dir
        self.base_name = base_name
        self.chunk_rows = max(1, int(chunk_rows))
        self.compression = compression
        self.min_free_bytes = min_free_bytes
        self.type_hints = type_hints or {}
        self.schema = _pin_schema(schema, self.type_hints) if schema else None

        self.paths: List[str] = []
        self._writer: Optional["pq.ParquetWriter"] = None
        self._current_path: Optional[str] = None
        self._rows_in_part = 0
        self._part_index = 0

        Path(output_dir).mkdir(parents=True, exist_ok=True)

    def __enter__(self) -> "ParquetPartWriter":
        return self

    def __exit__(self, exc_type, exc, tb) -> bool:
        if exc_type is not None:
            self.abort()
        else:
            self.close()
        return False

    def write(self, data: Any) -> None:
        """Write one batch: a pyarrow Table/RecordBatch or a Polars DataFrame."""
        table = _as_table(data)
        if self.schema is None:
            self.schema = _pin_schema(table.schema, self.type_hints)
        table = _conform(table, self.schema)

        offset = 0
        total = table.num_rows
        while offset < total:
            if self._writer is None:
                self._open_part()
            room = self.chunk_rows - self._rows_in_part
            take = min(room, total - offset)
            self._writer.write_table(
                table.slice(offset, take), row_group_size=self.chunk_rows
            )
            offset += take
            self._rows_in_part += take
            if self._rows_in_part >= self.chunk_rows:
                self._close_part()

    def close(self) -> List[str]:
        """Finish the object and return its parts, in order.

        An object that yielded no row still gets one empty part when its schema
        is known: downstream every object is expected to be scannable, and an
        empty table is a legitimate answer whereas a missing file is a crash.
        """
        self._close_part()
        if not self.paths and self.schema is not None:
            self._open_part()
            self._writer.write_table(self.schema.empty_table())
            self._close_part()
        return list(self.paths)

    def abort(self) -> None:
        """Close the current part after a failure, leaving what was written."""
        try:
            self._close_part()
        except Exception:  # noqa: BLE001 - the original error is the story
            logger.exception(
                "failed to close parquet part for %s", self.base_name
            )

    def _open_part(self) -> None:
        check_disk_space(self.output_dir, min_free_bytes=self.min_free_bytes)
        self._part_index += 1
        path = part_path(self.output_dir, self.base_name, self._part_index)
        self._writer = pq.ParquetWriter(
            path, self.schema, compression=self.compression
        )
        self._current_path = path
        self._rows_in_part = 0

    def _close_part(self) -> None:
        if self._writer is None:
            return
        self._writer.close()
        self._writer = None
        self.paths.append(self._current_path)
        self._rows_in_part = 0


# ---------------------------------------------------------------------------
# Producers -> parts
# ---------------------------------------------------------------------------


def write_arrow_batches(
    batches: Iterable[Any],
    output_dir: str,
    base_name: str,
    *,
    chunk_rows: int = DEFAULT_CHUNK_ROWS,
    schema: Optional["pa.Schema"] = None,
    type_hints: Optional[Dict[str, "pa.DataType"]] = None,
    min_free_bytes: int = DEFAULT_MIN_FREE_BYTES,
) -> List[str]:
    """Drain an iterator of Arrow batches into part files."""
    writer = ParquetPartWriter(
        output_dir,
        base_name,
        schema=schema,
        type_hints=type_hints,
        chunk_rows=chunk_rows,
        min_free_bytes=min_free_bytes,
    )
    try:
        for batch in batches:
            if batch is None:
                continue
            writer.write(batch)
    except Exception:
        writer.abort()
        raise
    return writer.close()


def write_row_batches(
    rows: Iterable[Sequence[Any]],
    columns: Sequence[str],
    output_dir: str,
    base_name: str,
    *,
    chunk_rows: int = DEFAULT_CHUNK_ROWS,
    fetch_rows: Optional[int] = None,
    type_hints: Optional[Dict[str, "pa.DataType"]] = None,
    min_free_bytes: int = DEFAULT_MIN_FREE_BYTES,
) -> List[str]:
    """Drain a DB-API style row iterator into part files, column-major.

    Row-major buffering (one Python dict per row, then a DataFrame per chunk) is
    what made the cursor-driven drivers allocate ``chunk_rows`` dicts per batch.
    Building one list per column and handing them to Arrow costs one list per
    column instead, and lets the schema be pinned on the first batch. The Python
    buffer holds ``fetch_rows`` rows, not a whole part: part size is a layout
    decision and should not set the memory bill.
    """
    names = list(columns)
    batch_rows = _batch_rows(chunk_rows, fetch_rows)
    writer = ParquetPartWriter(
        output_dir,
        base_name,
        type_hints=type_hints,
        chunk_rows=chunk_rows,
        min_free_bytes=min_free_bytes,
    )
    try:
        buffer: List[List[Any]] = [[] for _ in names]
        buffered = 0
        for row in rows:
            for index, value in enumerate(row):
                if index < len(buffer):
                    buffer[index].append(_arrow_safe(value))
            buffered += 1
            if buffered >= batch_rows:
                writer.write(_columns_to_table(names, buffer, writer.schema))
                buffer = [[] for _ in names]
                buffered = 0
        if buffered:
            writer.write(_columns_to_table(names, buffer, writer.schema))
    except Exception:
        writer.abort()
        raise
    return writer.close()


def write_dict_rows(
    documents: Iterable[Dict[str, Any]],
    output_dir: str,
    base_name: str,
    *,
    chunk_rows: int = DEFAULT_CHUNK_ROWS,
    fetch_rows: Optional[int] = None,
    min_free_bytes: int = DEFAULT_MIN_FREE_BYTES,
) -> List[str]:
    """Drain schemaless documents (MongoDB, Elasticsearch) into part files.

    The column set is fixed by the first batch. Keys that only appear later are
    dropped with a warning rather than silently widening part N and not part 1 —
    that widening is the schema drift that makes the parts unreadable together.
    """
    batch_rows = _batch_rows(chunk_rows, fetch_rows)
    writer = ParquetPartWriter(
        output_dir,
        base_name,
        chunk_rows=chunk_rows,
        min_free_bytes=min_free_bytes,
    )
    names: Optional[List[str]] = None
    dropped: set = set()
    try:
        buffer: List[Dict[str, Any]] = []
        for document in documents:
            buffer.append(document)
            if len(buffer) < batch_rows:
                continue
            names = names or _union_keys(buffer)
            _note_unknown_keys(buffer, names, dropped, base_name)
            writer.write(_documents_to_table(names, buffer, writer.schema))
            buffer = []
        if buffer:
            names = names or _union_keys(buffer)
            _note_unknown_keys(buffer, names, dropped, base_name)
            writer.write(_documents_to_table(names, buffer, writer.schema))
    except Exception:
        writer.abort()
        raise
    return writer.close()


def _union_keys(documents: Sequence[Dict[str, Any]]) -> List[str]:
    names: List[str] = []
    seen = set()
    for document in documents:
        for key in document:
            if key not in seen:
                seen.add(key)
                names.append(str(key))
    return names


def _note_unknown_keys(
    documents: Sequence[Dict[str, Any]],
    names: Sequence[str],
    dropped: set,
    base_name: str,
) -> None:
    known = set(names)
    for document in documents:
        for key in document:
            if key not in known and key not in dropped:
                dropped.add(key)
                logger.warning(
                    "%s: field %r appears after the first chunk and is not "
                    "part of the pinned schema; it is not exported",
                    base_name,
                    key,
                )


def _arrow_safe(value: Any) -> Any:
    """Make a driver value representable in a stable Arrow column.

    Nested documents and arrays are the reason a per-chunk inferred schema
    drifts; JSON is a representation Arrow can type identically in every part.
    """
    if isinstance(value, (dict, list, tuple, set)):
        return json.dumps(value, default=str)
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, (str, bytes, bool, int, float, type(None))):
        return value
    if isinstance(value, (_dt.datetime, _dt.date, _dt.time)):
        return value
    return str(value)


def _columns_to_table(
    names: Sequence[str],
    columns: Sequence[Sequence[Any]],
    schema: Optional["pa.Schema"],
) -> "pa.Table":
    arrays = []
    for index, name in enumerate(names):
        target = (
            schema.field(name).type
            if schema is not None and name in schema.names
            else None
        )
        arrays.append(_safe_array(columns[index], target))
    return pa.Table.from_arrays(arrays, names=list(names))


def _documents_to_table(
    names: Sequence[str],
    documents: Sequence[Dict[str, Any]],
    schema: Optional["pa.Schema"],
) -> "pa.Table":
    columns = [
        [_arrow_safe(document.get(name)) for document in documents]
        for name in names
    ]
    return _columns_to_table(names, columns, schema)


def _safe_array(
    values: Sequence[Any], target: Optional["pa.DataType"]
) -> "pa.Array":
    """Build an Arrow array, falling back to text when the values are mixed.

    A driver that returns an int for some rows and a string for others would
    otherwise abort the whole load; representing that column as text keeps every
    part of the object identically typed, which is what makes them scannable.
    """
    errors = (pa.ArrowInvalid, pa.ArrowTypeError, pa.ArrowNotImplementedError)
    if target is not None:
        try:
            return pa.array(values, type=target)
        except errors:
            pass
    try:
        return pa.array(values)
    except errors:
        return pa.array(
            [None if v is None else str(v) for v in values],
            type=pa.large_string(),
        )


# ---------------------------------------------------------------------------
# SQL
# ---------------------------------------------------------------------------


def stream_sql_to_parquet(
    engine: Any,
    sql: str,
    output_dir: str,
    base_name: str,
    *,
    chunk_rows: int = DEFAULT_CHUNK_ROWS,
    fetch_rows: Optional[int] = None,
    type_hints: Optional[Dict[str, "pa.DataType"]] = None,
    min_free_bytes: int = DEFAULT_MIN_FREE_BYTES,
) -> List[str]:
    """Stream a SQL result set to Parquet parts.

    ``stream_results``/``yield_per`` ask the DBAPI driver for a server-side
    cursor, which is the only way the client does not buffer the whole result
    set first. ``pandas.read_sql(..., chunksize=N)`` chunks client-side only:
    psycopg2, pymysql and pymssql have already materialized every row before
    pandas yields the first chunk, so a 100 GiB table dies inside libpq without
    Python seeing a row.

    Not every driver honours it — sqlite3, pymssql and most warehouse HTTP
    drivers ignore the option — and for those this is no worse than before while
    still writing Arrow batches straight to disk instead of building DataFrames.

    ``fetch_rows`` is deliberately separate from ``chunk_rows``: how many rows
    are alive at once is a memory question, how many rows go in a file is a
    layout question, and tying them together taxes the first for the second.
    Fetching 10k rows into 100k-row parts measured at a third of the peak RSS of
    fetching 100k.
    """
    check_disk_space(output_dir, min_free_bytes=min_free_bytes)
    batch_rows = _batch_rows(chunk_rows, fetch_rows)

    with engine.connect().execution_options(
        stream_results=True, yield_per=batch_rows
    ) as connection:
        batches = pl.read_database(
            sql,
            connection,
            iter_batches=True,
            batch_size=batch_rows,
            # Infer over the whole batch rather than its first 100 rows: a
            # column that is null in the first rows and populated later in the
            # same batch would otherwise be typed as Null.
            infer_schema_length=None,
        )
        writer = ParquetPartWriter(
            output_dir,
            base_name,
            type_hints=type_hints,
            chunk_rows=chunk_rows,
            min_free_bytes=min_free_bytes,
        )
        try:
            for batch in batches:
                writer.write(batch)
        except Exception:
            writer.abort()
            raise
        if writer.schema is None and type_hints:
            # An empty result set yields no batch at all, so the schema can only
            # come from the metadata; write the empty object rather than nothing.
            writer.schema = pa.schema(
                [pa.field(n, t) for n, t in type_hints.items()]
            )
        return writer.close()


# ---------------------------------------------------------------------------
# Files
# ---------------------------------------------------------------------------


def scan_csv(
    file_path: str,
    *,
    skip_rows: int = 0,
    encoding: str = "utf8",
    infer_schema_length: int = 10000,
    ignore_errors: bool = True,
    **kwargs: Any,
) -> "pl.LazyFrame":
    """Lazily scan a CSV. Nothing is read until the plan is executed."""
    if encoding.lower() in ("utf-8", "utf_8"):
        encoding = "utf8"
    return pl.scan_csv(
        file_path,
        skip_rows=skip_rows,
        encoding=encoding,
        infer_schema_length=infer_schema_length,
        ignore_errors=ignore_errors,
        **kwargs,
    )


def scan_ndjson(file_path: str, **kwargs: Any) -> "pl.LazyFrame":
    """Lazily scan newline-delimited JSON."""
    return pl.scan_ndjson(file_path, **kwargs)


def scan_parquet(source: Any, **kwargs: Any) -> "pl.LazyFrame":
    """Lazily scan Parquet file(s), local or remote."""
    return pl.scan_parquet(source, **kwargs)


def stream_to_parquet(
    lf: "pl.LazyFrame",
    output_path: str,
    *,
    row_group_size: int = DEFAULT_CHUNK_ROWS,
    compression: str = PARQUET_COMPRESSION,
    **kwargs: Any,
) -> str:
    """Sink a LazyFrame to one Parquet file without materializing it."""
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    lf.sink_parquet(
        output_path,
        row_group_size=row_group_size,
        compression=compression,
        **kwargs,
    )
    return output_path


def sink_parts(
    lf: "pl.LazyFrame",
    output_dir: str,
    base_name: str,
    *,
    chunk_rows: int = DEFAULT_CHUNK_ROWS,
    compression: str = PARQUET_COMPRESSION,
    min_free_bytes: int = DEFAULT_MIN_FREE_BYTES,
) -> List[str]:
    """Sink a LazyFrame to ``<base>_part_<n>.parquet`` in a SINGLE pass.

    The partitioned sink splits inside the engine. Slicing the plan per part
    instead — what ``stream_to_parquet_chunks`` did — re-executes the whole plan
    once per slice, so writing N parts meant reading the source N times.
    """
    check_disk_space(output_dir, min_free_bytes=min_free_bytes)
    Path(output_dir).mkdir(parents=True, exist_ok=True)

    written: List[str] = []

    def _name(context: Any) -> str:
        index = getattr(context, "index_in_partition", len(written))
        name = f"{base_name}_part_{int(index) + 1}.parquet"
        written.append(os.path.join(output_dir, name))
        return name

    lf.sink_parquet(
        _partitioning(output_dir, _name, chunk_rows),
        compression=compression,
        row_group_size=chunk_rows,
    )

    paths = sorted(p for p in written if os.path.exists(p))
    if paths:
        return _sorted_parts(paths)

    # No row at all: keep the object scannable by writing its (empty) schema.
    return [
        stream_to_parquet(
            lf.head(0),
            part_path(output_dir, base_name, 1),
            row_group_size=chunk_rows,
            compression=compression,
        )
    ]


def _partitioning(output_dir: str, namer: Any, chunk_rows: int) -> Any:
    """Partitioned-sink target, across the Polars versions we support."""
    if hasattr(pl, "PartitionBy"):
        return pl.PartitionBy(
            output_dir,
            file_path_provider=namer,
            max_rows_per_file=int(chunk_rows),
            # Let the row count alone decide the split, so part files line up
            # with chunk_rows instead of with an estimated in-memory size.
            approximate_bytes_per_file=None,
        )
    return pl.PartitionMaxSize(
        output_dir, file_path=namer, max_size=int(chunk_rows)
    )


def _sorted_parts(paths: Sequence[str]) -> List[str]:
    def _index(path: str) -> int:
        stem = Path(path).stem
        try:
            return int(stem.rsplit("_part_", 1)[1])
        except (IndexError, ValueError):
            return 0

    return sorted(paths, key=_index)


def sniff_json_format(file_path: str) -> str:
    """Return ``"array"`` or ``"ndjson"`` by looking at the first token.

    The loader used to take this from a ``json_lines`` config flag that defaults
    to False, so an NDJSON file was read with ``pl.read_json`` — the whole
    document at once — unless the caller happened to know to set it.
    """
    with open(file_path, "r", encoding="utf-8", errors="replace") as handle:
        while True:
            character = handle.read(1)
            if not character:
                return "ndjson"
            if not character.isspace():
                return "array" if character == "[" else "ndjson"


def iter_json_array(
    file_path: str, *, buffer_size: int = 1 << 20
) -> Iterator[Any]:
    """Yield the elements of a top-level JSON array one at a time.

    ``json.load`` builds the whole document plus its Python object graph before
    returning; a 2 GiB array of records is tens of GiB of dicts. This keeps one
    element and one buffer alive at a time, which is all a batching writer
    needs. It deliberately supports only the shape the loader can map to a
    table: a top-level array.
    """
    decoder = json.JSONDecoder()
    with open(file_path, "r", encoding="utf-8") as handle:
        buffer = handle.read(buffer_size)
        index = _skip_space(buffer, 0)
        if index >= len(buffer) or buffer[index] != "[":
            raise ValueError(
                f"{file_path} does not start with a JSON array; only "
                f"newline-delimited JSON and top-level arrays can be streamed."
            )
        index += 1

        while True:
            buffer = buffer[index:]
            index = 0
            while True:
                index = _skip_space(buffer, index)
                if index < len(buffer):
                    break
                more = handle.read(buffer_size)
                if not more:
                    return
                buffer += more

            if buffer[index] == ",":
                index += 1
                continue
            if buffer[index] == "]":
                return

            while True:
                try:
                    value, index = decoder.raw_decode(buffer, index)
                    break
                except ValueError:
                    more = handle.read(buffer_size)
                    if not more:
                        raise
                    buffer += more
            yield value


def _skip_space(text: str, index: int) -> int:
    while index < len(text) and text[index].isspace():
        index += 1
    return index


def iter_excel_rows(
    file_path: str,
    *,
    skip_rows: int = 0,
    sheet_name: Optional[str] = None,
):
    """Yield ``(headers, row_iterator)`` for a worksheet, row by row.

    Neither Polars nor pandas can stream an xlsx: both build the whole sheet
    before returning it. openpyxl's read-only mode is the only iterator over the
    file, so it is the only path the loader uses for Excel.
    """
    from openpyxl import load_workbook

    workbook = load_workbook(
        filename=file_path, read_only=True, data_only=True
    )
    worksheet = workbook[sheet_name] if sheet_name else workbook.active
    if worksheet is None:
        raise ValueError(f"{file_path} has no readable worksheet")

    rows = worksheet.iter_rows(values_only=True)
    for _ in range(int(skip_rows)):
        try:
            next(rows)
        except StopIteration:
            break
    try:
        headers = [
            str(name) if name is not None else f"column_{index + 1}"
            for index, name in enumerate(next(rows))
        ]
    except StopIteration:
        headers = []

    return headers, rows
