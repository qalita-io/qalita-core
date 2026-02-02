"""
# QALITA (c) COPYRIGHT 2025 - ALL RIGHTS RESERVED -
"""

import os
import glob
import pandas as pd
from typing import Optional, List, Iterable
from sqlalchemy import create_engine, inspect, text
from abc import ABC, abstractmethod
from pathlib import Path
from qalita_core.utils import slugify

DEFAULT_PORTS = {
    "5432": "postgresql",
    "3306": "mysql",
    "1433": "mssql+pymssql",  # Also used by Azure Synapse
    "1521": "oracle",
    "27017": "mongodb",
    "6379": "redis",
    "50000": "db2",
    "1434": "sybase",
    "3307": "mariadb",
    "5433": "greenplum",
    "5000": "sqlite",
    # Data warehouse ports
    "443": "snowflake",  # Snowflake uses HTTPS
    "5439": "redshift",
    "8123": "clickhouse",  # ClickHouse HTTP
    "9000": "clickhouse_native",  # ClickHouse native
    "8080": "trino",
    # Additional database ports
    "1025": "teradata",
    "30015": "sap_hana",
    "9042": "cassandra",
    "9200": "elasticsearch",
    "50001": "ibm_db2",
    "444": "athena",  # Athena uses HTTPS
}


class DataSource(ABC):
    @abstractmethod
    def get_data(self, table_or_query=None, pack_config=None):
        """Return a list of parquet file paths for the requested data."""
        pass


# -----------------------------
# Helper utilities for Parquet
# -----------------------------


def _ensure_output_dir(pack_config: Optional[dict]) -> str:
    base_dir = (pack_config or {}).get("parquet_output_dir") or "./parquet"
    Path(base_dir).mkdir(parents=True, exist_ok=True)
    return str(base_dir)


def _build_base_name(source_type: str, object_identifier: str) -> str:
    normalized_source = slugify(source_type or "source")
    normalized_object = slugify(object_identifier or "data")
    return f"{normalized_source}_{normalized_object}"


def _build_parquet_path(output_dir: str, base_name: str, part_index: int) -> str:
    return os.path.join(output_dir, f"{base_name}_part_{part_index}.parquet")


def cleanup_parquet_files(paths: List[str], logger=None) -> int:
    """
    Remove parquet files from the filesystem.

    Args:
        paths: List of parquet file paths to remove.
        logger: Optional logger for debug/error messages.

    Returns:
        Number of files successfully removed.
    """
    removed_count = 0
    for path in paths or []:
        try:
            if os.path.exists(path):
                os.remove(path)
                removed_count += 1
                if logger:
                    logger.debug(f"Removed temporary parquet file: {path}")
        except OSError as e:
            if logger:
                logger.warning(f"Failed to remove parquet file {path}: {e}")
    return removed_count


def _write_df_to_parquet(df: pd.DataFrame, output_path: str) -> str:
    # Always write with pyarrow for Arrow/Polars/DuckDB compatibility
    df.to_parquet(output_path, engine="pyarrow", index=False)
    return output_path


def _write_pandas_chunks(
    df_iter: Iterable[pd.DataFrame],
    output_dir: str,
    base_name: str,
    start_part: int = 1,
) -> List[str]:
    paths: List[str] = []
    part = start_part
    for chunk_df in df_iter:
        path = _build_parquet_path(output_dir, base_name, part)
        _write_df_to_parquet(chunk_df, path)
        paths.append(path)
        part += 1
    return paths


class FileSource(DataSource):
    def __init__(self, file_path):
        self.file_path = file_path

    def get_data(self, table_or_query=None, pack_config=None):
        output_dir = _ensure_output_dir(pack_config)
        if os.path.isfile(self.file_path):
            return self._load_file(self.file_path, pack_config, output_dir)
        if os.path.isdir(self.file_path):
            data_files = glob.glob(os.path.join(self.file_path, "*.csv")) + glob.glob(
                os.path.join(self.file_path, "*.xlsx")
            )
            if not data_files:
                raise FileNotFoundError(
                    "No CSV or XLSX files found in the provided path."
                )
            return self._load_file(data_files[0], pack_config, output_dir)
        raise FileNotFoundError(
            f"The path {self.file_path} is neither a file nor a directory, or it can't be reached."
        )

    @staticmethod
    def _load_file(file_path, pack_config, output_dir: str) -> List[str]:
        skiprows = 0
        chunk_rows = (pack_config or {}).get("chunk_rows", 100000)
        if pack_config:
            skiprows = pack_config.get("job", {}).get("source", {}).get("skiprows", 0)

        base_name = _build_base_name(
            "file", os.path.splitext(os.path.basename(file_path))[0]
        )
        # CSV: stream with chunksize
        if file_path.endswith(".csv"):
            df_iter = pd.read_csv(
                file_path,
                low_memory=False,
                memory_map=True,
                skiprows=int(skiprows),
                on_bad_lines="warn",
                encoding="utf-8",
                chunksize=int(chunk_rows),
            )
            return _write_pandas_chunks(df_iter, output_dir, base_name)

        # XLSX: stream via openpyxl read_only
        if file_path.endswith(".xlsx"):
            try:
                from openpyxl import load_workbook
            except Exception:
                # Fallback: load entire file (may be memory heavy)
                df = pd.read_excel(file_path, engine="openpyxl", skiprows=int(skiprows))
                path = _build_parquet_path(output_dir, base_name, 1)
                return [_write_df_to_parquet(df, path)]

            wb = load_workbook(filename=file_path, read_only=True, data_only=True)
            ws = wb.active
            if ws is None:
                # Fallback: load entire file
                df = pd.read_excel(file_path, engine="openpyxl", skiprows=int(skiprows))
                path = _build_parquet_path(output_dir, base_name, 1)
                return [_write_df_to_parquet(df, path)]

            rows_iter = ws.iter_rows(values_only=True)
            # Apply skiprows on header discovery
            for _ in range(int(skiprows)):
                try:
                    next(rows_iter)
                except StopIteration:
                    break
            try:
                headers = list(next(rows_iter))
            except StopIteration:
                headers = []

            batch: List[list] = []
            part = 1
            paths: List[str] = []
            for row in rows_iter:
                batch.append(list(row))
                if len(batch) >= int(chunk_rows):
                    df = pd.DataFrame(batch, columns=headers)
                    path = _build_parquet_path(output_dir, base_name, part)
                    _write_df_to_parquet(df, path)
                    paths.append(path)
                    batch = []
                    part += 1
            if batch:
                df = pd.DataFrame(batch, columns=headers)
                path = _build_parquet_path(output_dir, base_name, part)
                _write_df_to_parquet(df, path)
                paths.append(path)
            return paths

        raise ValueError(
            f"Unsupported file extension or missing 'skiprows' for file: {file_path}"
        )


class DatabaseSource(DataSource):
    def __init__(self, connection_string=None, config=None):
        # Keep the config available for schema preference and other options
        self.config = config or {}

        if connection_string:
            self.engine = create_engine(connection_string)
        elif config:
            db_type = config.get("type") or DEFAULT_PORTS.get(
                str(config.get("port")), "unknown"
            )
            if db_type == "unknown":
                raise ValueError(
                    f"Unsupported or unknown database port: {config.get('port')}"
                )
            elif db_type == "oracle":
                db_type = "oracle+oracledb"
                conn_str = (
                    f"{db_type}://{config['username']}:{config['password']}"
                    f"@{config['host']}:{config['port']}/?service_name={config['database']}"
                )
                self.engine = create_engine(conn_str)
            elif db_type.startswith("sqlite"):
                database_path = config.get("database") or ":memory:"
                if database_path == ":memory:":
                    conn_str = "sqlite:///:memory:"
                else:
                    # Accept absolute or relative filesystem path
                    conn_str = f"sqlite:///{database_path}"
                self.engine = create_engine(conn_str)
            else:
                self.engine = create_engine(
                    f"{db_type}://{config['username']}:{config['password']}@{config['host']}:{config['port']}/{config['database']}"
                )
        else:
            raise ValueError(
                "DatabaseSource requires a connection_string or a config dict."
            )

    def get_data(self, table_or_query=None, pack_config=None):
        """
        Load data from the database and write Parquet chunks.

        - If table_or_query is a string table name: returns list of parquet paths for that table
        - If table_or_query is a SQL query string: returns list of parquet paths for the query result
        - If table_or_query is a list/tuple/set of table names: returns parquet paths for each table
        - If table_or_query is '*' or None: scan all tables and return parquet paths for each
        """

        # Determine schema: prefer source config over pack config; both are optional
        schema = None
        cfg_schema = None
        try:
            cfg_schema = (self.config or {}).get("schema")
        except Exception:
            cfg_schema = None
        if cfg_schema:
            schema = cfg_schema
        elif pack_config:
            schema = pack_config.get("job", {}).get("source", {}).get("schema")

        output_dir = _ensure_output_dir(pack_config)
        chunk_rows = int((pack_config or {}).get("chunk_rows", 100000))
        dialect_name = None
        try:
            dialect_name = self.engine.dialect.name
        except Exception:
            dialect_name = None

        # Default behavior: scan all tables
        if table_or_query is None or (
            isinstance(table_or_query, str) and table_or_query.strip() == "*"
        ):
            table_names = self._get_all_table_names(schema)
            if not table_names:
                raise ValueError(
                    "No tables found in the database for the given schema."
                )
            all_paths: List[str] = []
            for table_name in table_names:
                all_paths.extend(
                    self._read_table_to_parquet(
                        table_name, schema, output_dir, chunk_rows, dialect_name
                    )
                )
            return all_paths

        # If a list/tuple/set of table names is provided
        if isinstance(table_or_query, (list, tuple, set)):
            table_names = list(table_or_query)
            all_paths: List[str] = []
            for table_name in table_names:
                all_paths.extend(
                    self._read_table_to_parquet(
                        table_name, schema, output_dir, chunk_rows, dialect_name
                    )
                )
            return all_paths

        # If a single string is provided, determine if it's a table name or SQL query
        if isinstance(table_or_query, str):
            if self._is_sql_query(table_or_query):
                base_name = _build_base_name(dialect_name or "db", "query")
                df_iter = pd.read_sql(table_or_query, self.engine, chunksize=chunk_rows)
                return _write_pandas_chunks(df_iter, output_dir, base_name)
            return self._read_table_to_parquet(
                table_or_query, schema, output_dir, chunk_rows, dialect_name
            )

        raise TypeError(
            "table_or_query must be None, '*', a string (table name or SQL), or a list/tuple/set of table names."
        )

    def _read_table(
        self, table_name: str, schema: Optional[str] = None
    ) -> pd.DataFrame:
        """Read a full table as a DataFrame using dialect-aware SQL. (Compatibility only)"""
        # Support fully-qualified names like "SCHEMA.TABLE" when no schema is explicitly provided
        effective_schema = schema
        effective_table = table_name
        if not schema and "." in table_name:
            try:
                effective_schema, effective_table = table_name.split(".", 1)
            except ValueError:
                effective_schema = None
                effective_table = table_name

        try:
            return pd.read_sql_table(
                effective_table, self.engine, schema=effective_schema
            )
        except Exception:
            # Fallback to a simple SELECT * if read_sql_table is unsupported for the dialect
            qualified = (
                f"{effective_schema}.{effective_table}"
                if effective_schema
                else effective_table
            )
            return pd.read_sql(f"SELECT * FROM {qualified}", self.engine)

    def _read_table_to_parquet(
        self,
        table_name: str,
        schema: Optional[str],
        output_dir: str,
        chunk_rows: int,
        dialect_name: Optional[str],
    ) -> List[str]:
        effective_schema = schema
        effective_table = table_name
        if not schema and "." in table_name:
            try:
                effective_schema, effective_table = table_name.split(".", 1)
            except ValueError:
                effective_schema = None
                effective_table = table_name

        qualified = (
            f"{effective_schema}.{effective_table}"
            if effective_schema
            else effective_table
        )
        base_name = _build_base_name(dialect_name or "db", qualified)
        # Use streaming SQL with chunksize
        sql = f"SELECT * FROM {qualified}"
        df_iter = pd.read_sql(sql, self.engine, chunksize=int(chunk_rows))
        return _write_pandas_chunks(df_iter, output_dir, base_name)

    def _get_all_table_names(self, schema: Optional[str] = None) -> List[str]:
        """Return all table names (and views) in the database for the given schema, sorted alphabetically.
        For Oracle and PostgreSQL, if no schema is provided and none are found in the default schema,
        iterate over accessible schemas and return fully-qualified names ("SCHEMA.TABLE").
        """
        inspector = inspect(self.engine)

        def _collect_for_schema(target_schema: Optional[str]) -> List[str]:
            try:
                tables = inspector.get_table_names(schema=target_schema)
            except Exception:
                tables = []
            try:
                views = inspector.get_view_names(schema=target_schema)
            except Exception:
                views = []
            return list(set((tables or []) + (views or [])))

        # First pass: use the provided schema (or None)
        initial = sorted(_collect_for_schema(schema))
        if initial:
            return initial

        # Special handling for Oracle when no schema specified and nothing found
        try:
            dialect_name = self.engine.dialect.name
        except Exception:
            dialect_name = None

        if dialect_name == "oracle" and schema is None:
            try:
                schemas = inspector.get_schema_names()
            except Exception:
                schemas = []

            system_schemas = {
                "SYS",
                "SYSTEM",
                "OUTLN",
                "XDB",
                "MDSYS",
                "CTXSYS",
                "ORDSYS",
                "ORDDATA",
                "DBSNMP",
                "APPQOSSYS",
                "WMSYS",
                "OLAPSYS",
                "LBACSYS",
                "GSMADMIN_INTERNAL",
                "OJVMSYS",
                "DVF",
                "DVSYS",
                "REMOTE_SCHEDULER_AGENT",
                "SYS$UMF",
                "GGSYS",
                "AUDSYS",
                "ANONYMOUS",
            }

            qualified: List[str] = []
            for sch in schemas or []:
                if not sch or sch.upper() in system_schemas:
                    continue
                names = _collect_for_schema(sch)
                for n in names:
                    qualified.append(f"{sch}.{n}")

            qualified = sorted(set(qualified))
            if qualified:
                return qualified

            # Final fallback: try CURRENT_SCHEMA if available
            try:
                with self.engine.connect() as conn:
                    result = conn.execute(
                        text("SELECT sys_context('USERENV','CURRENT_SCHEMA') FROM dual")
                    )
                    row = result.fetchone()
                    current_schema = row[0] if row and row[0] else None
                if current_schema:
                    names = _collect_for_schema(current_schema)
                    if names:
                        return sorted([f"{current_schema}.{n}" for n in names])
            except Exception:
                pass

        # Special handling for PostgreSQL when no schema specified and nothing found
        if dialect_name == "postgresql" and schema is None:
            try:
                schemas = inspector.get_schema_names()
            except Exception:
                schemas = []

            # Filter out system schemas
            system_prefixes = ("pg_temp_", "pg_toast_temp_")
            system_schemas = {
                "information_schema",
                "pg_catalog",
                "pg_toast",
            }

            qualified: List[str] = []
            for sch in schemas or []:
                if not sch:
                    continue
                if sch in system_schemas or any(
                    sch.startswith(pfx) for pfx in system_prefixes
                ):
                    continue
                names = _collect_for_schema(sch)
                for n in names:
                    qualified.append(f"{sch}.{n}")

            qualified = sorted(set(qualified))
            if qualified:
                return qualified

        # Return whatever was found initially (likely empty)
        return initial

    def _is_sql_query(self, s: str) -> bool:
        """Heuristic to detect if a string is a SQL query rather than a bare table name."""
        sql = s.strip().lower()
        if ";" in sql or "\n" in sql:
            return True
        starters = ("select", "with", "show", "describe", "pragma", "explain")
        return any(sql.startswith(token) for token in starters)


def _infer_format_from_path(path: str, explicit_format: Optional[str] = None) -> str:
    if explicit_format:
        return explicit_format.lower()
    lower = path.lower()
    if lower.endswith(".csv"):
        return "csv"
    if lower.endswith(".json"):
        return "json"
    if lower.endswith(".parquet") or lower.endswith(".pq"):
        return "parquet"
    if lower.endswith(".xlsx") or lower.endswith(".xls"):
        return "excel"
    return "csv"


def _materialize_remote_to_parquet(
    path: str,
    fmt: str,
    storage_options: Optional[dict],
    pack_config: Optional[dict],
) -> List[str]:
    # If already parquet, pass-through by returning the remote path
    if fmt == "parquet":
        return [path]

    output_dir = _ensure_output_dir(pack_config)
    chunk_rows = int((pack_config or {}).get("chunk_rows", 100000))
    skiprows = 0
    if pack_config:
        skiprows = pack_config.get("job", {}).get("source", {}).get("skiprows", 0)

    base_name = _build_base_name("remote", os.path.splitext(os.path.basename(path))[0])

    if fmt == "csv":
        df_iter = pd.read_csv(
            path,
            storage_options=storage_options,
            low_memory=False,
            memory_map=True,
            skiprows=int(skiprows),
            on_bad_lines="warn",
            encoding="utf-8",
            chunksize=chunk_rows,
        )
        return _write_pandas_chunks(df_iter, output_dir, base_name)
    if fmt == "json":
        # Attempt newline-delimited JSON if hinted
        lines = bool((pack_config or {}).get("json_lines", False))
        if lines:
            df_iter = pd.read_json(
                path, storage_options=storage_options, lines=True, chunksize=chunk_rows
            )
            return _write_pandas_chunks(df_iter, output_dir, base_name)
        # Fallback: load once, write once (may be memory heavy)
        df = pd.read_json(path, storage_options=storage_options)
        return [_write_df_to_parquet(df, _build_parquet_path(output_dir, base_name, 1))]
    if fmt == "excel":
        # Excel streaming from remote is complex; fallback to single load
        df = pd.read_excel(
            path,
            storage_options=storage_options,
            engine="openpyxl",
            skiprows=int(skiprows),
        )
        return [_write_df_to_parquet(df, _build_parquet_path(output_dir, base_name, 1))]
    # Fallback to CSV behavior
    df_iter = pd.read_csv(path, storage_options=storage_options, chunksize=chunk_rows)
    return _write_pandas_chunks(df_iter, output_dir, base_name)


class S3Source(DataSource):
    def __init__(self, config):
        self.config = config or {}

    def get_data(self, table_or_query=None, pack_config=None):
        # Expect either a full s3 path in config['path'] or bucket/key
        path = self.config.get("path")
        if not path:
            bucket = self.config.get("bucket")
            key = self.config.get("key")
            if bucket and key:
                path = f"s3://{bucket}/{key}"
        if not path:
            raise ValueError(
                "S3Source requires either 'path' or 'bucket'+'key' in config."
            )

        storage_options = {}
        for opt_key in [
            "key",  # aws_access_key_id
            "secret",  # aws_secret_access_key
            "token",  # aws_session_token
            "client_kwargs",  # e.g., {"region_name": "us-east-1"}
        ]:
            if opt_key in self.config:
                storage_options[opt_key] = self.config[opt_key]

        fmt = _infer_format_from_path(path, self.config.get("format"))
        return _materialize_remote_to_parquet(
            path, fmt, storage_options or None, pack_config
        )


class GCSSource(DataSource):
    def __init__(self, config):
        self.config = config or {}

    def get_data(self, table_or_query=None, pack_config=None):
        # Expect gs:// style path or bucket/object
        path = self.config.get("path")
        if not path:
            bucket = self.config.get("bucket")
            blob = self.config.get("blob") or self.config.get("key")
            if bucket and blob:
                path = f"gs://{bucket}/{blob}"
        if not path:
            raise ValueError(
                "GCSSource requires either 'path' or 'bucket'+'blob' in config."
            )

        storage_options = {}
        for opt_key in [
            "token",  # path to service account json or dict credentials
            "project",
        ]:
            if opt_key in self.config:
                storage_options[opt_key] = self.config[opt_key]

        fmt = _infer_format_from_path(path, self.config.get("format"))
        return _materialize_remote_to_parquet(
            path, fmt, storage_options or None, pack_config
        )


class AzureBlobSource(DataSource):
    def __init__(self, config):
        self.config = config or {}

    def get_data(self, table_or_query=None, pack_config=None):
        # Accept full abfs(s):// path or account/container/blob components
        path = self.config.get("path")
        if not path:
            account_name = self.config.get("account_name")
            container = self.config.get("container")
            blob = self.config.get("blob") or self.config.get("key")
            if account_name and container and blob:
                path = f"abfs://{container}@{account_name}.dfs.core.windows.net/{blob}"
        if not path:
            raise ValueError(
                "AzureBlobSource requires either 'path' or 'account_name'+'container'+'blob'."
            )

        storage_options = {}
        # adlfs uses Azure credentials via storage_options
        for opt_key in [
            "account_name",
            "account_key",
            "sas_token",
            "tenant_id",
            "client_id",
            "client_secret",
        ]:
            if opt_key in self.config:
                storage_options[opt_key] = self.config[opt_key]

        fmt = _infer_format_from_path(path, self.config.get("format"))
        return _materialize_remote_to_parquet(
            path, fmt, storage_options or None, pack_config
        )


class HDFSSource(DataSource):
    def __init__(self, config):
        self.config = config or {}

    def get_data(self, table_or_query=None, pack_config=None):
        # Expect hdfs://host:port/path
        path = self.config.get("path")
        if not path:
            host = self.config.get("host")
            port = self.config.get("port") or 8020
            hdfs_path = self.config.get("hdfs_path") or self.config.get("key")
            if host and hdfs_path:
                path = f"hdfs://{host}:{port}/{hdfs_path.lstrip('/')}"
        if not path:
            raise ValueError(
                "HDFSSource requires 'path' or 'host'+'hdfs_path' in config."
            )

        storage_options = {}
        for opt_key in [
            "host",
            "port",
            "user",
            "kerb_kwargs",  # kerberos parameters if applicable
        ]:
            if opt_key in self.config:
                storage_options[opt_key] = self.config[opt_key]

        fmt = _infer_format_from_path(path, self.config.get("format"))
        return _materialize_remote_to_parquet(
            path, fmt, storage_options or None, pack_config
        )


class FolderSource(DataSource):
    def __init__(self, config):
        self.config = config

    def get_data(self, table_or_query=None, pack_config=None):
        raise NotImplementedError("FolderSource.get_data Not yet Implemented.")


class MongoDBSource(DataSource):
    """MongoDB data source using pymongo."""

    def __init__(self, config):
        self.config = config or {}

    def get_data(self, table_or_query=None, pack_config=None):
        """
        Load data from MongoDB and write Parquet chunks.

        - If table_or_query is a string collection name: returns parquet paths for that collection
        - If table_or_query is a list of collection names: returns parquet paths for each
        - If table_or_query is '*' or None: scan all collections
        """
        try:
            from pymongo import MongoClient
        except ImportError:
            raise ImportError("pymongo is required for MongoDBSource. Install with: pip install pymongo")

        output_dir = _ensure_output_dir(pack_config)
        chunk_rows = int((pack_config or {}).get("chunk_rows", 100000))

        # Build connection
        connection_string = self.config.get("connection_string")
        if connection_string:
            client = MongoClient(connection_string)
        else:
            host = self.config.get("host", "localhost")
            port = int(self.config.get("port", 27017))
            username = self.config.get("username")
            password = self.config.get("password")
            database = self.config.get("database")

            if username and password:
                client = MongoClient(
                    host=host,
                    port=port,
                    username=username,
                    password=password,
                    authSource=database or "admin",
                )
            else:
                client = MongoClient(host=host, port=port)

        database_name = self.config.get("database")
        if not database_name:
            raise ValueError("MongoDBSource requires 'database' in config.")

        db = client[database_name]

        # Determine collections to process
        if table_or_query is None or (isinstance(table_or_query, str) and table_or_query.strip() == "*"):
            collections = db.list_collection_names()
        elif isinstance(table_or_query, (list, tuple, set)):
            collections = list(table_or_query)
        elif isinstance(table_or_query, str):
            collections = [table_or_query]
        else:
            raise TypeError("table_or_query must be None, '*', a string, or a list of collection names.")

        all_paths: List[str] = []
        for collection_name in collections:
            collection = db[collection_name]
            base_name = _build_base_name("mongodb", collection_name)

            # Stream documents in batches
            cursor = collection.find()
            batch: List[dict] = []
            part = 1

            for doc in cursor:
                # Convert ObjectId to string for serialization
                if "_id" in doc:
                    doc["_id"] = str(doc["_id"])
                batch.append(doc)

                if len(batch) >= chunk_rows:
                    df = pd.DataFrame(batch)
                    path = _build_parquet_path(output_dir, base_name, part)
                    _write_df_to_parquet(df, path)
                    all_paths.append(path)
                    batch = []
                    part += 1

            # Write remaining documents
            if batch:
                df = pd.DataFrame(batch)
                path = _build_parquet_path(output_dir, base_name, part)
                _write_df_to_parquet(df, path)
                all_paths.append(path)

        client.close()
        return all_paths


class SnowflakeSource(DataSource):
    """Snowflake data warehouse source using snowflake-sqlalchemy."""

    def __init__(self, config):
        self.config = config or {}

    def get_data(self, table_or_query=None, pack_config=None):
        """
        Load data from Snowflake and write Parquet chunks.
        """
        try:
            from sqlalchemy import create_engine
        except ImportError:
            raise ImportError("sqlalchemy is required for SnowflakeSource.")

        output_dir = _ensure_output_dir(pack_config)
        chunk_rows = int((pack_config or {}).get("chunk_rows", 100000))

        # Build connection string
        connection_string = self.config.get("connection_string")
        if not connection_string:
            account = self.config.get("account")
            user = self.config.get("user") or self.config.get("username")
            password = self.config.get("password")
            database = self.config.get("database")
            schema = self.config.get("schema", "PUBLIC")
            warehouse = self.config.get("warehouse")
            role = self.config.get("role")

            if not all([account, user, password]):
                raise ValueError("SnowflakeSource requires 'account', 'user', and 'password' in config.")

            # snowflake://user:password@account/database/schema?warehouse=WH&role=ROLE
            connection_string = f"snowflake://{user}:{password}@{account}"
            if database:
                connection_string += f"/{database}"
                if schema:
                    connection_string += f"/{schema}"

            params = []
            if warehouse:
                params.append(f"warehouse={warehouse}")
            if role:
                params.append(f"role={role}")
            if params:
                connection_string += "?" + "&".join(params)

        engine = create_engine(connection_string)
        schema = self.config.get("schema", "PUBLIC")

        return self._load_data(engine, table_or_query, schema, output_dir, chunk_rows, "snowflake")

    def _load_data(self, engine, table_or_query, schema, output_dir, chunk_rows, dialect_name):
        """Common data loading logic for SQL-based sources."""
        if table_or_query is None or (isinstance(table_or_query, str) and table_or_query.strip() == "*"):
            inspector = inspect(engine)
            table_names = inspector.get_table_names(schema=schema)
            all_paths: List[str] = []
            for table_name in table_names:
                all_paths.extend(
                    self._read_table_to_parquet(engine, table_name, schema, output_dir, chunk_rows, dialect_name)
                )
            return all_paths

        if isinstance(table_or_query, (list, tuple, set)):
            all_paths: List[str] = []
            for table_name in table_or_query:
                all_paths.extend(
                    self._read_table_to_parquet(engine, table_name, schema, output_dir, chunk_rows, dialect_name)
                )
            return all_paths

        if isinstance(table_or_query, str):
            if self._is_sql_query(table_or_query):
                base_name = _build_base_name(dialect_name, "query")
                df_iter = pd.read_sql(table_or_query, engine, chunksize=chunk_rows)
                return _write_pandas_chunks(df_iter, output_dir, base_name)
            return self._read_table_to_parquet(engine, table_or_query, schema, output_dir, chunk_rows, dialect_name)

        raise TypeError("table_or_query must be None, '*', a string, or a list of table names.")

    def _read_table_to_parquet(self, engine, table_name, schema, output_dir, chunk_rows, dialect_name):
        qualified = f"{schema}.{table_name}" if schema else table_name
        base_name = _build_base_name(dialect_name, qualified)
        sql = f"SELECT * FROM {qualified}"
        df_iter = pd.read_sql(sql, engine, chunksize=int(chunk_rows))
        return _write_pandas_chunks(df_iter, output_dir, base_name)

    def _is_sql_query(self, s: str) -> bool:
        sql = s.strip().lower()
        if ";" in sql or "\n" in sql:
            return True
        starters = ("select", "with", "show", "describe", "pragma", "explain")
        return any(sql.startswith(token) for token in starters)


class BigQuerySource(DataSource):
    """Google BigQuery data source using sqlalchemy-bigquery."""

    def __init__(self, config):
        self.config = config or {}

    def get_data(self, table_or_query=None, pack_config=None):
        """
        Load data from BigQuery and write Parquet chunks.
        """
        try:
            from sqlalchemy import create_engine
        except ImportError:
            raise ImportError("sqlalchemy is required for BigQuerySource.")

        output_dir = _ensure_output_dir(pack_config)
        chunk_rows = int((pack_config or {}).get("chunk_rows", 100000))

        # Build connection string
        connection_string = self.config.get("connection_string")
        if not connection_string:
            project = self.config.get("project")
            dataset = self.config.get("dataset")
            credentials_path = self.config.get("credentials_json") or self.config.get("credentials")

            if not project:
                raise ValueError("BigQuerySource requires 'project' in config.")

            # bigquery://project/dataset
            connection_string = f"bigquery://{project}"
            if dataset:
                connection_string += f"/{dataset}"

            if credentials_path:
                connection_string += f"?credentials_path={credentials_path}"

        engine = create_engine(connection_string)
        dataset = self.config.get("dataset")

        # Use similar pattern to SnowflakeSource
        if table_or_query is None or (isinstance(table_or_query, str) and table_or_query.strip() == "*"):
            inspector = inspect(engine)
            try:
                table_names = inspector.get_table_names(schema=dataset)
            except Exception:
                table_names = []
            all_paths: List[str] = []
            for table_name in table_names:
                all_paths.extend(
                    self._read_table_to_parquet(engine, table_name, dataset, output_dir, chunk_rows)
                )
            return all_paths

        if isinstance(table_or_query, (list, tuple, set)):
            all_paths: List[str] = []
            for table_name in table_or_query:
                all_paths.extend(
                    self._read_table_to_parquet(engine, table_name, dataset, output_dir, chunk_rows)
                )
            return all_paths

        if isinstance(table_or_query, str):
            if self._is_sql_query(table_or_query):
                base_name = _build_base_name("bigquery", "query")
                df_iter = pd.read_sql(table_or_query, engine, chunksize=chunk_rows)
                return _write_pandas_chunks(df_iter, output_dir, base_name)
            return self._read_table_to_parquet(engine, table_or_query, dataset, output_dir, chunk_rows)

        raise TypeError("table_or_query must be None, '*', a string, or a list of table names.")

    def _read_table_to_parquet(self, engine, table_name, dataset, output_dir, chunk_rows):
        qualified = f"{dataset}.{table_name}" if dataset else table_name
        base_name = _build_base_name("bigquery", qualified)
        sql = f"SELECT * FROM `{qualified}`"
        df_iter = pd.read_sql(sql, engine, chunksize=int(chunk_rows))
        return _write_pandas_chunks(df_iter, output_dir, base_name)

    def _is_sql_query(self, s: str) -> bool:
        sql = s.strip().lower()
        if ";" in sql or "\n" in sql:
            return True
        starters = ("select", "with", "show", "describe", "explain")
        return any(sql.startswith(token) for token in starters)


class DatabricksSource(DataSource):
    """Databricks data source using databricks-sql-connector."""

    def __init__(self, config):
        self.config = config or {}

    def get_data(self, table_or_query=None, pack_config=None):
        """
        Load data from Databricks and write Parquet chunks.
        """
        try:
            from databricks import sql as databricks_sql
        except ImportError:
            raise ImportError("databricks-sql-connector is required for DatabricksSource. Install with: pip install databricks-sql-connector")

        output_dir = _ensure_output_dir(pack_config)
        chunk_rows = int((pack_config or {}).get("chunk_rows", 100000))

        server_hostname = self.config.get("server_hostname") or self.config.get("host")
        http_path = self.config.get("http_path")
        access_token = self.config.get("access_token") or self.config.get("token")
        catalog = self.config.get("catalog")
        schema = self.config.get("schema")

        if not all([server_hostname, http_path, access_token]):
            raise ValueError("DatabricksSource requires 'server_hostname', 'http_path', and 'access_token' in config.")

        connection = databricks_sql.connect(
            server_hostname=server_hostname,
            http_path=http_path,
            access_token=access_token,
        )

        cursor = connection.cursor()

        # Set catalog and schema if provided
        if catalog:
            cursor.execute(f"USE CATALOG {catalog}")
        if schema:
            cursor.execute(f"USE SCHEMA {schema}")

        # Determine tables to process
        if table_or_query is None or (isinstance(table_or_query, str) and table_or_query.strip() == "*"):
            cursor.execute("SHOW TABLES")
            tables = [row[1] for row in cursor.fetchall()]  # table name is typically second column
        elif isinstance(table_or_query, (list, tuple, set)):
            tables = list(table_or_query)
        elif isinstance(table_or_query, str):
            if self._is_sql_query(table_or_query):
                # Execute query directly
                base_name = _build_base_name("databricks", "query")
                cursor.execute(table_or_query)
                columns = [desc[0] for desc in cursor.description]
                all_paths: List[str] = []
                batch = []
                part = 1
                for row in cursor:
                    batch.append(dict(zip(columns, row)))
                    if len(batch) >= chunk_rows:
                        df = pd.DataFrame(batch)
                        path = _build_parquet_path(output_dir, base_name, part)
                        _write_df_to_parquet(df, path)
                        all_paths.append(path)
                        batch = []
                        part += 1
                if batch:
                    df = pd.DataFrame(batch)
                    path = _build_parquet_path(output_dir, base_name, part)
                    _write_df_to_parquet(df, path)
                    all_paths.append(path)
                cursor.close()
                connection.close()
                return all_paths
            tables = [table_or_query]
        else:
            raise TypeError("table_or_query must be None, '*', a string, or a list of table names.")

        all_paths: List[str] = []
        for table_name in tables:
            base_name = _build_base_name("databricks", table_name)
            cursor.execute(f"SELECT * FROM {table_name}")
            columns = [desc[0] for desc in cursor.description]
            batch = []
            part = 1
            for row in cursor:
                batch.append(dict(zip(columns, row)))
                if len(batch) >= chunk_rows:
                    df = pd.DataFrame(batch)
                    path = _build_parquet_path(output_dir, base_name, part)
                    _write_df_to_parquet(df, path)
                    all_paths.append(path)
                    batch = []
                    part += 1
            if batch:
                df = pd.DataFrame(batch)
                path = _build_parquet_path(output_dir, base_name, part)
                _write_df_to_parquet(df, path)
                all_paths.append(path)

        cursor.close()
        connection.close()
        return all_paths

    def _is_sql_query(self, s: str) -> bool:
        sql = s.strip().lower()
        if ";" in sql or "\n" in sql:
            return True
        starters = ("select", "with", "show", "describe", "explain")
        return any(sql.startswith(token) for token in starters)


class RedshiftSource(DataSource):
    """Amazon Redshift data source using redshift-connector or psycopg2."""

    def __init__(self, config):
        self.config = config or {}

    def get_data(self, table_or_query=None, pack_config=None):
        """
        Load data from Redshift and write Parquet chunks.
        """
        output_dir = _ensure_output_dir(pack_config)
        chunk_rows = int((pack_config or {}).get("chunk_rows", 100000))

        # Build connection string - Redshift is PostgreSQL-compatible
        connection_string = self.config.get("connection_string")
        if not connection_string:
            host = self.config.get("host")
            port = self.config.get("port", 5439)
            user = self.config.get("user") or self.config.get("username")
            password = self.config.get("password")
            database = self.config.get("database")

            if not all([host, user, password, database]):
                raise ValueError("RedshiftSource requires 'host', 'user', 'password', and 'database' in config.")

            # Try redshift+redshift_connector first, fall back to postgresql
            try:
                import redshift_connector  # noqa: F401
                connection_string = f"redshift+redshift_connector://{user}:{password}@{host}:{port}/{database}"
            except ImportError:
                # Fall back to PostgreSQL driver
                connection_string = f"postgresql://{user}:{password}@{host}:{port}/{database}"

        engine = create_engine(connection_string)
        schema = self.config.get("schema", "public")

        # Use DatabaseSource pattern
        db_source = DatabaseSource(connection_string=connection_string, config=self.config)
        return db_source.get_data(table_or_query=table_or_query, pack_config=pack_config)


class ClickHouseSource(DataSource):
    """ClickHouse data source using clickhouse-sqlalchemy."""

    def __init__(self, config):
        self.config = config or {}

    def get_data(self, table_or_query=None, pack_config=None):
        """
        Load data from ClickHouse and write Parquet chunks.
        """
        output_dir = _ensure_output_dir(pack_config)
        chunk_rows = int((pack_config or {}).get("chunk_rows", 100000))

        # Build connection string
        connection_string = self.config.get("connection_string")
        if not connection_string:
            host = self.config.get("host", "localhost")
            port = self.config.get("port", 8123)
            user = self.config.get("user") or self.config.get("username", "default")
            password = self.config.get("password", "")
            database = self.config.get("database", "default")
            protocol = self.config.get("protocol", "http")  # http or native

            # clickhouse+http://user:password@host:port/database
            # clickhouse+native://user:password@host:port/database
            driver = f"clickhouse+{protocol}"
            if password:
                connection_string = f"{driver}://{user}:{password}@{host}:{port}/{database}"
            else:
                connection_string = f"{driver}://{user}@{host}:{port}/{database}"

        engine = create_engine(connection_string)
        database = self.config.get("database", "default")

        # Get tables
        if table_or_query is None or (isinstance(table_or_query, str) and table_or_query.strip() == "*"):
            inspector = inspect(engine)
            try:
                table_names = inspector.get_table_names()
            except Exception:
                # Fallback: query system tables
                with engine.connect() as conn:
                    result = conn.execute(text(f"SHOW TABLES FROM {database}"))
                    table_names = [row[0] for row in result]
            all_paths: List[str] = []
            for table_name in table_names:
                all_paths.extend(
                    self._read_table_to_parquet(engine, table_name, output_dir, chunk_rows)
                )
            return all_paths

        if isinstance(table_or_query, (list, tuple, set)):
            all_paths: List[str] = []
            for table_name in table_or_query:
                all_paths.extend(
                    self._read_table_to_parquet(engine, table_name, output_dir, chunk_rows)
                )
            return all_paths

        if isinstance(table_or_query, str):
            if self._is_sql_query(table_or_query):
                base_name = _build_base_name("clickhouse", "query")
                df_iter = pd.read_sql(table_or_query, engine, chunksize=chunk_rows)
                return _write_pandas_chunks(df_iter, output_dir, base_name)
            return self._read_table_to_parquet(engine, table_or_query, output_dir, chunk_rows)

        raise TypeError("table_or_query must be None, '*', a string, or a list of table names.")

    def _read_table_to_parquet(self, engine, table_name, output_dir, chunk_rows):
        base_name = _build_base_name("clickhouse", table_name)
        sql = f"SELECT * FROM {table_name}"
        df_iter = pd.read_sql(sql, engine, chunksize=int(chunk_rows))
        return _write_pandas_chunks(df_iter, output_dir, base_name)

    def _is_sql_query(self, s: str) -> bool:
        sql = s.strip().lower()
        if ";" in sql or "\n" in sql:
            return True
        starters = ("select", "with", "show", "describe", "explain")
        return any(sql.startswith(token) for token in starters)


class DuckDBSource(DataSource):
    """DuckDB data source (local files or MotherDuck cloud)."""

    def __init__(self, config):
        self.config = config or {}

    def get_data(self, table_or_query=None, pack_config=None):
        """
        Load data from DuckDB and write Parquet chunks.
        """
        try:
            import duckdb
        except ImportError:
            raise ImportError("duckdb is required for DuckDBSource. Install with: pip install duckdb")

        output_dir = _ensure_output_dir(pack_config)
        chunk_rows = int((pack_config or {}).get("chunk_rows", 100000))

        # Connect to DuckDB
        path = self.config.get("path") or self.config.get("database", ":memory:")
        motherduck_token = self.config.get("motherduck_token") or self.config.get("token")

        if motherduck_token:
            # MotherDuck cloud connection
            connection_string = f"md:{path}?motherduck_token={motherduck_token}"
            conn = duckdb.connect(connection_string)
        else:
            conn = duckdb.connect(path)

        schema = self.config.get("schema", "main")

        # Get tables
        if table_or_query is None or (isinstance(table_or_query, str) and table_or_query.strip() == "*"):
            result = conn.execute(f"SELECT table_name FROM information_schema.tables WHERE table_schema = '{schema}'")
            table_names = [row[0] for row in result.fetchall()]
        elif isinstance(table_or_query, (list, tuple, set)):
            table_names = list(table_or_query)
        elif isinstance(table_or_query, str):
            if self._is_sql_query(table_or_query):
                base_name = _build_base_name("duckdb", "query")
                result = conn.execute(table_or_query)
                df = result.df()
                all_paths: List[str] = []
                # Split into chunks
                for i in range(0, len(df), chunk_rows):
                    chunk_df = df.iloc[i:i + chunk_rows]
                    path = _build_parquet_path(output_dir, base_name, (i // chunk_rows) + 1)
                    _write_df_to_parquet(chunk_df, path)
                    all_paths.append(path)
                conn.close()
                return all_paths if all_paths else [_write_df_to_parquet(df, _build_parquet_path(output_dir, base_name, 1))]
            table_names = [table_or_query]
        else:
            raise TypeError("table_or_query must be None, '*', a string, or a list of table names.")

        all_paths: List[str] = []
        for table_name in table_names:
            base_name = _build_base_name("duckdb", table_name)
            qualified = f"{schema}.{table_name}" if schema != "main" else table_name
            result = conn.execute(f"SELECT * FROM {qualified}")
            df = result.df()
            # Split into chunks
            if len(df) == 0:
                path = _build_parquet_path(output_dir, base_name, 1)
                _write_df_to_parquet(df, path)
                all_paths.append(path)
            else:
                for i in range(0, len(df), chunk_rows):
                    chunk_df = df.iloc[i:i + chunk_rows]
                    path = _build_parquet_path(output_dir, base_name, (i // chunk_rows) + 1)
                    _write_df_to_parquet(chunk_df, path)
                    all_paths.append(path)

        conn.close()
        return all_paths

    def _is_sql_query(self, s: str) -> bool:
        sql = s.strip().lower()
        if ";" in sql or "\n" in sql:
            return True
        starters = ("select", "with", "show", "describe", "explain", "pragma")
        return any(sql.startswith(token) for token in starters)


class TrinoSource(DataSource):
    """Trino/Presto data source using trino library."""

    def __init__(self, config):
        self.config = config or {}

    def get_data(self, table_or_query=None, pack_config=None):
        """
        Load data from Trino and write Parquet chunks.
        """
        try:
            from trino.dbapi import connect as trino_connect
        except ImportError:
            raise ImportError("trino is required for TrinoSource. Install with: pip install trino")

        output_dir = _ensure_output_dir(pack_config)
        chunk_rows = int((pack_config or {}).get("chunk_rows", 100000))

        host = self.config.get("host", "localhost")
        port = int(self.config.get("port", 8080))
        user = self.config.get("user") or self.config.get("username", "trino")
        catalog = self.config.get("catalog")
        schema = self.config.get("schema")
        http_scheme = self.config.get("http_scheme", "https")

        conn = trino_connect(
            host=host,
            port=port,
            user=user,
            catalog=catalog,
            schema=schema,
            http_scheme=http_scheme,
        )
        cursor = conn.cursor()

        # Get tables
        if table_or_query is None or (isinstance(table_or_query, str) and table_or_query.strip() == "*"):
            cursor.execute("SHOW TABLES")
            table_names = [row[0] for row in cursor.fetchall()]
        elif isinstance(table_or_query, (list, tuple, set)):
            table_names = list(table_or_query)
        elif isinstance(table_or_query, str):
            if self._is_sql_query(table_or_query):
                base_name = _build_base_name("trino", "query")
                cursor.execute(table_or_query)
                columns = [desc[0] for desc in cursor.description]
                all_paths: List[str] = []
                batch = []
                part = 1
                for row in cursor:
                    batch.append(dict(zip(columns, row)))
                    if len(batch) >= chunk_rows:
                        df = pd.DataFrame(batch)
                        path = _build_parquet_path(output_dir, base_name, part)
                        _write_df_to_parquet(df, path)
                        all_paths.append(path)
                        batch = []
                        part += 1
                if batch:
                    df = pd.DataFrame(batch)
                    path = _build_parquet_path(output_dir, base_name, part)
                    _write_df_to_parquet(df, path)
                    all_paths.append(path)
                cursor.close()
                conn.close()
                return all_paths
            table_names = [table_or_query]
        else:
            raise TypeError("table_or_query must be None, '*', a string, or a list of table names.")

        all_paths: List[str] = []
        for table_name in table_names:
            base_name = _build_base_name("trino", table_name)
            cursor.execute(f"SELECT * FROM {table_name}")
            columns = [desc[0] for desc in cursor.description]
            batch = []
            part = 1
            for row in cursor:
                batch.append(dict(zip(columns, row)))
                if len(batch) >= chunk_rows:
                    df = pd.DataFrame(batch)
                    path = _build_parquet_path(output_dir, base_name, part)
                    _write_df_to_parquet(df, path)
                    all_paths.append(path)
                    batch = []
                    part += 1
            if batch:
                df = pd.DataFrame(batch)
                path = _build_parquet_path(output_dir, base_name, part)
                _write_df_to_parquet(df, path)
                all_paths.append(path)

        cursor.close()
        conn.close()
        return all_paths

    def _is_sql_query(self, s: str) -> bool:
        sql = s.strip().lower()
        if ";" in sql or "\n" in sql:
            return True
        starters = ("select", "with", "show", "describe", "explain")
        return any(sql.startswith(token) for token in starters)


class TeradataSource(DataSource):
    """Teradata data source using teradatasqlalchemy."""

    def __init__(self, config):
        self.config = config or {}

    def get_data(self, table_or_query=None, pack_config=None):
        """Load data from Teradata and write Parquet chunks."""
        output_dir = _ensure_output_dir(pack_config)
        chunk_rows = int((pack_config or {}).get("chunk_rows", 100000))

        connection_string = self.config.get("connection_string")
        if not connection_string:
            host = self.config.get("host")
            user = self.config.get("user") or self.config.get("username")
            password = self.config.get("password")
            database = self.config.get("database")

            if not all([host, user, password]):
                raise ValueError("TeradataSource requires 'host', 'user', and 'password' in config.")

            # teradatasql://user:password@host/database
            connection_string = f"teradatasql://{user}:{password}@{host}"
            if database:
                connection_string += f"/{database}"

        engine = create_engine(connection_string)
        schema = self.config.get("schema")

        return self._load_data(engine, table_or_query, schema, output_dir, chunk_rows, "teradata")

    def _load_data(self, engine, table_or_query, schema, output_dir, chunk_rows, dialect_name):
        if table_or_query is None or (isinstance(table_or_query, str) and table_or_query.strip() == "*"):
            inspector = inspect(engine)
            table_names = inspector.get_table_names(schema=schema)
            all_paths: List[str] = []
            for table_name in table_names:
                all_paths.extend(
                    self._read_table_to_parquet(engine, table_name, schema, output_dir, chunk_rows, dialect_name)
                )
            return all_paths

        if isinstance(table_or_query, (list, tuple, set)):
            all_paths: List[str] = []
            for table_name in table_or_query:
                all_paths.extend(
                    self._read_table_to_parquet(engine, table_name, schema, output_dir, chunk_rows, dialect_name)
                )
            return all_paths

        if isinstance(table_or_query, str):
            if self._is_sql_query(table_or_query):
                base_name = _build_base_name(dialect_name, "query")
                df_iter = pd.read_sql(table_or_query, engine, chunksize=chunk_rows)
                return _write_pandas_chunks(df_iter, output_dir, base_name)
            return self._read_table_to_parquet(engine, table_or_query, schema, output_dir, chunk_rows, dialect_name)

        raise TypeError("table_or_query must be None, '*', a string, or a list of table names.")

    def _read_table_to_parquet(self, engine, table_name, schema, output_dir, chunk_rows, dialect_name):
        qualified = f"{schema}.{table_name}" if schema else table_name
        base_name = _build_base_name(dialect_name, qualified)
        sql = f"SELECT * FROM {qualified}"
        df_iter = pd.read_sql(sql, engine, chunksize=int(chunk_rows))
        return _write_pandas_chunks(df_iter, output_dir, base_name)

    def _is_sql_query(self, s: str) -> bool:
        sql = s.strip().lower()
        if ";" in sql or "\n" in sql:
            return True
        starters = ("select", "with", "show", "describe", "explain")
        return any(sql.startswith(token) for token in starters)


class SapHanaSource(DataSource):
    """SAP HANA data source using hdbcli/sqlalchemy-hana."""

    def __init__(self, config):
        self.config = config or {}

    def get_data(self, table_or_query=None, pack_config=None):
        """Load data from SAP HANA and write Parquet chunks."""
        output_dir = _ensure_output_dir(pack_config)
        chunk_rows = int((pack_config or {}).get("chunk_rows", 100000))

        connection_string = self.config.get("connection_string")
        if not connection_string:
            host = self.config.get("host")
            port = self.config.get("port", 30015)
            user = self.config.get("user") or self.config.get("username")
            password = self.config.get("password")
            database = self.config.get("database")

            if not all([host, user, password]):
                raise ValueError("SapHanaSource requires 'host', 'user', and 'password' in config.")

            # hana://user:password@host:port
            connection_string = f"hana://{user}:{password}@{host}:{port}"

        engine = create_engine(connection_string)
        schema = self.config.get("schema")

        return self._load_data(engine, table_or_query, schema, output_dir, chunk_rows, "sap_hana")

    def _load_data(self, engine, table_or_query, schema, output_dir, chunk_rows, dialect_name):
        if table_or_query is None or (isinstance(table_or_query, str) and table_or_query.strip() == "*"):
            inspector = inspect(engine)
            table_names = inspector.get_table_names(schema=schema)
            all_paths: List[str] = []
            for table_name in table_names:
                all_paths.extend(
                    self._read_table_to_parquet(engine, table_name, schema, output_dir, chunk_rows, dialect_name)
                )
            return all_paths

        if isinstance(table_or_query, (list, tuple, set)):
            all_paths: List[str] = []
            for table_name in table_or_query:
                all_paths.extend(
                    self._read_table_to_parquet(engine, table_name, schema, output_dir, chunk_rows, dialect_name)
                )
            return all_paths

        if isinstance(table_or_query, str):
            if self._is_sql_query(table_or_query):
                base_name = _build_base_name(dialect_name, "query")
                df_iter = pd.read_sql(table_or_query, engine, chunksize=chunk_rows)
                return _write_pandas_chunks(df_iter, output_dir, base_name)
            return self._read_table_to_parquet(engine, table_or_query, schema, output_dir, chunk_rows, dialect_name)

        raise TypeError("table_or_query must be None, '*', a string, or a list of table names.")

    def _read_table_to_parquet(self, engine, table_name, schema, output_dir, chunk_rows, dialect_name):
        qualified = f'"{schema}"."{table_name}"' if schema else f'"{table_name}"'
        base_name = _build_base_name(dialect_name, f"{schema}.{table_name}" if schema else table_name)
        sql = f"SELECT * FROM {qualified}"
        df_iter = pd.read_sql(sql, engine, chunksize=int(chunk_rows))
        return _write_pandas_chunks(df_iter, output_dir, base_name)

    def _is_sql_query(self, s: str) -> bool:
        sql = s.strip().lower()
        if ";" in sql or "\n" in sql:
            return True
        starters = ("select", "with", "show", "describe", "explain")
        return any(sql.startswith(token) for token in starters)


class CassandraSource(DataSource):
    """Apache Cassandra data source using cassandra-driver."""

    def __init__(self, config):
        self.config = config or {}

    def get_data(self, table_or_query=None, pack_config=None):
        """Load data from Cassandra and write Parquet chunks."""
        try:
            from cassandra.cluster import Cluster
            from cassandra.auth import PlainTextAuthProvider
        except ImportError:
            raise ImportError("cassandra-driver is required for CassandraSource. Install with: pip install cassandra-driver")

        output_dir = _ensure_output_dir(pack_config)
        chunk_rows = int((pack_config or {}).get("chunk_rows", 100000))

        hosts = self.config.get("hosts") or [self.config.get("host", "localhost")]
        port = int(self.config.get("port", 9042))
        username = self.config.get("username")
        password = self.config.get("password")
        keyspace = self.config.get("keyspace") or self.config.get("database")

        if username and password:
            auth_provider = PlainTextAuthProvider(username=username, password=password)
            cluster = Cluster(hosts, port=port, auth_provider=auth_provider)
        else:
            cluster = Cluster(hosts, port=port)

        session = cluster.connect(keyspace)

        # Get tables
        if table_or_query is None or (isinstance(table_or_query, str) and table_or_query.strip() == "*"):
            rows = session.execute(f"SELECT table_name FROM system_schema.tables WHERE keyspace_name = '{keyspace}'")
            table_names = [row.table_name for row in rows]
        elif isinstance(table_or_query, (list, tuple, set)):
            table_names = list(table_or_query)
        elif isinstance(table_or_query, str):
            if self._is_sql_query(table_or_query):
                base_name = _build_base_name("cassandra", "query")
                rows = session.execute(table_or_query)
                columns = rows.column_names
                all_paths: List[str] = []
                batch = []
                part = 1
                for row in rows:
                    batch.append(dict(zip(columns, row)))
                    if len(batch) >= chunk_rows:
                        df = pd.DataFrame(batch)
                        path = _build_parquet_path(output_dir, base_name, part)
                        _write_df_to_parquet(df, path)
                        all_paths.append(path)
                        batch = []
                        part += 1
                if batch:
                    df = pd.DataFrame(batch)
                    path = _build_parquet_path(output_dir, base_name, part)
                    _write_df_to_parquet(df, path)
                    all_paths.append(path)
                cluster.shutdown()
                return all_paths
            table_names = [table_or_query]
        else:
            raise TypeError("table_or_query must be None, '*', a string, or a list of table names.")

        all_paths: List[str] = []
        for table_name in table_names:
            base_name = _build_base_name("cassandra", table_name)
            rows = session.execute(f"SELECT * FROM {table_name}")
            columns = rows.column_names
            batch = []
            part = 1
            for row in rows:
                batch.append(dict(zip(columns, row)))
                if len(batch) >= chunk_rows:
                    df = pd.DataFrame(batch)
                    path = _build_parquet_path(output_dir, base_name, part)
                    _write_df_to_parquet(df, path)
                    all_paths.append(path)
                    batch = []
                    part += 1
            if batch:
                df = pd.DataFrame(batch)
                path = _build_parquet_path(output_dir, base_name, part)
                _write_df_to_parquet(df, path)
                all_paths.append(path)

        cluster.shutdown()
        return all_paths

    def _is_sql_query(self, s: str) -> bool:
        sql = s.strip().lower()
        if ";" in sql or "\n" in sql:
            return True
        starters = ("select", "with")
        return any(sql.startswith(token) for token in starters)


class ElasticsearchSource(DataSource):
    """Elasticsearch data source."""

    def __init__(self, config):
        self.config = config or {}

    def get_data(self, table_or_query=None, pack_config=None):
        """Load data from Elasticsearch and write Parquet chunks."""
        try:
            from elasticsearch import Elasticsearch
        except ImportError:
            raise ImportError("elasticsearch is required for ElasticsearchSource. Install with: pip install elasticsearch")

        output_dir = _ensure_output_dir(pack_config)
        chunk_rows = int((pack_config or {}).get("chunk_rows", 100000))

        hosts = self.config.get("hosts") or [self.config.get("host", "localhost")]
        port = int(self.config.get("port", 9200))
        username = self.config.get("username")
        password = self.config.get("password")
        use_ssl = self.config.get("use_ssl", False)
        verify_certs = self.config.get("verify_certs", True)
        api_key = self.config.get("api_key")
        cloud_id = self.config.get("cloud_id")

        # Build hosts with port
        if isinstance(hosts, list) and hosts and not any(":" in str(h) for h in hosts):
            scheme = "https" if use_ssl else "http"
            hosts = [f"{scheme}://{h}:{port}" for h in hosts]

        if cloud_id:
            es = Elasticsearch(cloud_id=cloud_id, api_key=api_key) if api_key else Elasticsearch(cloud_id=cloud_id, basic_auth=(username, password))
        elif api_key:
            es = Elasticsearch(hosts=hosts, api_key=api_key, verify_certs=verify_certs)
        elif username and password:
            es = Elasticsearch(hosts=hosts, basic_auth=(username, password), verify_certs=verify_certs)
        else:
            es = Elasticsearch(hosts=hosts, verify_certs=verify_certs)

        # Get indices (table_or_query is the index pattern)
        if table_or_query is None or (isinstance(table_or_query, str) and table_or_query.strip() == "*"):
            indices = list(es.indices.get(index="*").keys())
            # Filter out system indices
            indices = [i for i in indices if not i.startswith(".")]
        elif isinstance(table_or_query, (list, tuple, set)):
            indices = list(table_or_query)
        elif isinstance(table_or_query, str):
            indices = [table_or_query]
        else:
            raise TypeError("table_or_query must be None, '*', a string (index name), or a list of index names.")

        all_paths: List[str] = []
        for index in indices:
            base_name = _build_base_name("elasticsearch", index)
            # Use scroll API for large datasets
            resp = es.search(index=index, scroll="2m", size=min(chunk_rows, 10000), body={"query": {"match_all": {}}})
            scroll_id = resp["_scroll_id"]
            hits = resp["hits"]["hits"]

            batch = []
            part = 1
            while hits:
                for hit in hits:
                    doc = hit["_source"]
                    doc["_id"] = hit["_id"]
                    batch.append(doc)
                    if len(batch) >= chunk_rows:
                        df = pd.DataFrame(batch)
                        path = _build_parquet_path(output_dir, base_name, part)
                        _write_df_to_parquet(df, path)
                        all_paths.append(path)
                        batch = []
                        part += 1
                resp = es.scroll(scroll_id=scroll_id, scroll="2m")
                scroll_id = resp["_scroll_id"]
                hits = resp["hits"]["hits"]

            if batch:
                df = pd.DataFrame(batch)
                path = _build_parquet_path(output_dir, base_name, part)
                _write_df_to_parquet(df, path)
                all_paths.append(path)

            es.clear_scroll(scroll_id=scroll_id)

        es.close()
        return all_paths


class IbmDb2Source(DataSource):
    """IBM DB2 data source using ibm_db_sa."""

    def __init__(self, config):
        self.config = config or {}

    def get_data(self, table_or_query=None, pack_config=None):
        """Load data from IBM DB2 and write Parquet chunks."""
        output_dir = _ensure_output_dir(pack_config)
        chunk_rows = int((pack_config or {}).get("chunk_rows", 100000))

        connection_string = self.config.get("connection_string")
        if not connection_string:
            host = self.config.get("host")
            port = self.config.get("port", 50000)
            user = self.config.get("user") or self.config.get("username")
            password = self.config.get("password")
            database = self.config.get("database")

            if not all([host, user, password, database]):
                raise ValueError("IbmDb2Source requires 'host', 'user', 'password', and 'database' in config.")

            # ibm_db_sa://user:password@host:port/database
            connection_string = f"ibm_db_sa://{user}:{password}@{host}:{port}/{database}"

        engine = create_engine(connection_string)
        schema = self.config.get("schema")

        return self._load_data(engine, table_or_query, schema, output_dir, chunk_rows, "ibm_db2")

    def _load_data(self, engine, table_or_query, schema, output_dir, chunk_rows, dialect_name):
        if table_or_query is None or (isinstance(table_or_query, str) and table_or_query.strip() == "*"):
            inspector = inspect(engine)
            table_names = inspector.get_table_names(schema=schema)
            all_paths: List[str] = []
            for table_name in table_names:
                all_paths.extend(
                    self._read_table_to_parquet(engine, table_name, schema, output_dir, chunk_rows, dialect_name)
                )
            return all_paths

        if isinstance(table_or_query, (list, tuple, set)):
            all_paths: List[str] = []
            for table_name in table_or_query:
                all_paths.extend(
                    self._read_table_to_parquet(engine, table_name, schema, output_dir, chunk_rows, dialect_name)
                )
            return all_paths

        if isinstance(table_or_query, str):
            if self._is_sql_query(table_or_query):
                base_name = _build_base_name(dialect_name, "query")
                df_iter = pd.read_sql(table_or_query, engine, chunksize=chunk_rows)
                return _write_pandas_chunks(df_iter, output_dir, base_name)
            return self._read_table_to_parquet(engine, table_or_query, schema, output_dir, chunk_rows, dialect_name)

        raise TypeError("table_or_query must be None, '*', a string, or a list of table names.")

    def _read_table_to_parquet(self, engine, table_name, schema, output_dir, chunk_rows, dialect_name):
        qualified = f"{schema}.{table_name}" if schema else table_name
        base_name = _build_base_name(dialect_name, qualified)
        sql = f"SELECT * FROM {qualified}"
        df_iter = pd.read_sql(sql, engine, chunksize=int(chunk_rows))
        return _write_pandas_chunks(df_iter, output_dir, base_name)

    def _is_sql_query(self, s: str) -> bool:
        sql = s.strip().lower()
        if ";" in sql or "\n" in sql:
            return True
        starters = ("select", "with", "show", "describe", "explain")
        return any(sql.startswith(token) for token in starters)


class AthenaSource(DataSource):
    """Amazon Athena data source using PyAthena."""

    def __init__(self, config):
        self.config = config or {}

    def get_data(self, table_or_query=None, pack_config=None):
        """Load data from Amazon Athena and write Parquet chunks."""
        output_dir = _ensure_output_dir(pack_config)
        chunk_rows = int((pack_config or {}).get("chunk_rows", 100000))

        connection_string = self.config.get("connection_string")
        if not connection_string:
            region = self.config.get("region", "us-east-1")
            s3_staging_dir = self.config.get("s3_staging_dir")
            database = self.config.get("database")
            access_key = self.config.get("access_key") or self.config.get("aws_access_key_id")
            secret_key = self.config.get("secret_key") or self.config.get("aws_secret_access_key")
            workgroup = self.config.get("workgroup")

            if not s3_staging_dir:
                raise ValueError("AthenaSource requires 's3_staging_dir' in config.")

            # awsathena+rest://access_key:secret_key@athena.region.amazonaws.com:443/database?s3_staging_dir=s3://...
            if access_key and secret_key:
                connection_string = f"awsathena+rest://{access_key}:{secret_key}@athena.{region}.amazonaws.com:443/"
            else:
                connection_string = f"awsathena+rest://:@athena.{region}.amazonaws.com:443/"

            if database:
                connection_string += database

            connection_string += f"?s3_staging_dir={s3_staging_dir}"
            if workgroup:
                connection_string += f"&work_group={workgroup}"

        engine = create_engine(connection_string)
        schema = self.config.get("schema") or self.config.get("database")

        return self._load_data(engine, table_or_query, schema, output_dir, chunk_rows, "athena")

    def _load_data(self, engine, table_or_query, schema, output_dir, chunk_rows, dialect_name):
        if table_or_query is None or (isinstance(table_or_query, str) and table_or_query.strip() == "*"):
            inspector = inspect(engine)
            try:
                table_names = inspector.get_table_names(schema=schema)
            except Exception:
                table_names = []
            all_paths: List[str] = []
            for table_name in table_names:
                all_paths.extend(
                    self._read_table_to_parquet(engine, table_name, schema, output_dir, chunk_rows, dialect_name)
                )
            return all_paths

        if isinstance(table_or_query, (list, tuple, set)):
            all_paths: List[str] = []
            for table_name in table_or_query:
                all_paths.extend(
                    self._read_table_to_parquet(engine, table_name, schema, output_dir, chunk_rows, dialect_name)
                )
            return all_paths

        if isinstance(table_or_query, str):
            if self._is_sql_query(table_or_query):
                base_name = _build_base_name(dialect_name, "query")
                df_iter = pd.read_sql(table_or_query, engine, chunksize=chunk_rows)
                return _write_pandas_chunks(df_iter, output_dir, base_name)
            return self._read_table_to_parquet(engine, table_or_query, schema, output_dir, chunk_rows, dialect_name)

        raise TypeError("table_or_query must be None, '*', a string, or a list of table names.")

    def _read_table_to_parquet(self, engine, table_name, schema, output_dir, chunk_rows, dialect_name):
        qualified = f'"{schema}"."{table_name}"' if schema else f'"{table_name}"'
        base_name = _build_base_name(dialect_name, f"{schema}.{table_name}" if schema else table_name)
        sql = f"SELECT * FROM {qualified}"
        df_iter = pd.read_sql(sql, engine, chunksize=int(chunk_rows))
        return _write_pandas_chunks(df_iter, output_dir, base_name)

    def _is_sql_query(self, s: str) -> bool:
        sql = s.strip().lower()
        if ";" in sql or "\n" in sql:
            return True
        starters = ("select", "with", "show", "describe", "explain")
        return any(sql.startswith(token) for token in starters)


class SynapseSource(DataSource):
    """Azure Synapse Analytics data source."""

    def __init__(self, config):
        self.config = config or {}

    def get_data(self, table_or_query=None, pack_config=None):
        """Load data from Azure Synapse and write Parquet chunks."""
        output_dir = _ensure_output_dir(pack_config)
        chunk_rows = int((pack_config or {}).get("chunk_rows", 100000))

        connection_string = self.config.get("connection_string")
        if not connection_string:
            host = self.config.get("host")  # xxx.sql.azuresynapse.net
            port = self.config.get("port", 1433)
            user = self.config.get("user") or self.config.get("username")
            password = self.config.get("password")
            database = self.config.get("database")

            if not all([host, user, password, database]):
                raise ValueError("SynapseSource requires 'host', 'user', 'password', and 'database' in config.")

            # mssql+pyodbc://user:password@host:port/database?driver=ODBC+Driver+17+for+SQL+Server
            # Or use pymssql: mssql+pymssql://user:password@host:port/database
            connection_string = f"mssql+pymssql://{user}:{password}@{host}:{port}/{database}"

        engine = create_engine(connection_string)
        schema = self.config.get("schema", "dbo")

        return self._load_data(engine, table_or_query, schema, output_dir, chunk_rows, "synapse")

    def _load_data(self, engine, table_or_query, schema, output_dir, chunk_rows, dialect_name):
        if table_or_query is None or (isinstance(table_or_query, str) and table_or_query.strip() == "*"):
            inspector = inspect(engine)
            table_names = inspector.get_table_names(schema=schema)
            all_paths: List[str] = []
            for table_name in table_names:
                all_paths.extend(
                    self._read_table_to_parquet(engine, table_name, schema, output_dir, chunk_rows, dialect_name)
                )
            return all_paths

        if isinstance(table_or_query, (list, tuple, set)):
            all_paths: List[str] = []
            for table_name in table_or_query:
                all_paths.extend(
                    self._read_table_to_parquet(engine, table_name, schema, output_dir, chunk_rows, dialect_name)
                )
            return all_paths

        if isinstance(table_or_query, str):
            if self._is_sql_query(table_or_query):
                base_name = _build_base_name(dialect_name, "query")
                df_iter = pd.read_sql(table_or_query, engine, chunksize=chunk_rows)
                return _write_pandas_chunks(df_iter, output_dir, base_name)
            return self._read_table_to_parquet(engine, table_or_query, schema, output_dir, chunk_rows, dialect_name)

        raise TypeError("table_or_query must be None, '*', a string, or a list of table names.")

    def _read_table_to_parquet(self, engine, table_name, schema, output_dir, chunk_rows, dialect_name):
        qualified = f"[{schema}].[{table_name}]" if schema else f"[{table_name}]"
        base_name = _build_base_name(dialect_name, f"{schema}.{table_name}" if schema else table_name)
        sql = f"SELECT * FROM {qualified}"
        df_iter = pd.read_sql(sql, engine, chunksize=int(chunk_rows))
        return _write_pandas_chunks(df_iter, output_dir, base_name)

    def _is_sql_query(self, s: str) -> bool:
        sql = s.strip().lower()
        if ";" in sql or "\n" in sql:
            return True
        starters = ("select", "with", "show", "describe", "explain")
        return any(sql.startswith(token) for token in starters)


class SqliteSource(DataSource):
    def __init__(self, config):
        self.config = config

    def get_data(self, table_or_query=None, pack_config=None):
        raise NotImplementedError("SqliteSource.get_data Not yet Implemented.")


def get_data_source(source_config):
    type_ = source_config.get("type")
    config = source_config.get("config", {})

    # File sources
    if type_ in ("file", "csv", "excel", "json", "parquet"):
        return FileSource(config.get("path"))
    elif type_ == "folder":
        return FolderSource(config)

    # Traditional relational databases (via SQLAlchemy)
    elif type_ == "postgresql":
        return DatabaseSource(
            connection_string=config.get("connection_string"),
            config=config,
        )
    elif type_ == "mysql":
        return DatabaseSource(
            connection_string=config.get("connection_string"),
            config=config,
        )
    elif type_ == "oracle":
        return DatabaseSource(
            connection_string=config.get("connection_string"),
            config=config,
        )
    elif type_ == "mssql":
        return DatabaseSource(
            connection_string=config.get("connection_string"),
            config=config,
        )
    elif type_ == "sqlite":
        return DatabaseSource(
            connection_string=config.get("connection_string"),
            config=config,
        )

    # NoSQL databases
    elif type_ == "mongodb":
        return MongoDBSource(config)

    # Cloud object storage
    elif type_ == "s3":
        return S3Source(config)
    elif type_ == "gcs":
        return GCSSource(config)
    elif type_ == "azure_blob":
        return AzureBlobSource(config)
    elif type_ == "hdfs":
        return HDFSSource(config)

    # Data warehouses
    elif type_ == "snowflake":
        return SnowflakeSource(config)
    elif type_ == "bigquery":
        return BigQuerySource(config)
    elif type_ == "databricks":
        return DatabricksSource(config)
    elif type_ == "redshift":
        return RedshiftSource(config)
    elif type_ == "clickhouse":
        return ClickHouseSource(config)
    elif type_ == "duckdb":
        return DuckDBSource(config)
    elif type_ == "trino":
        return TrinoSource(config)

    # Enterprise databases
    elif type_ == "teradata":
        return TeradataSource(config)
    elif type_ == "sap_hana":
        return SapHanaSource(config)
    elif type_ == "cassandra":
        return CassandraSource(config)
    elif type_ == "elasticsearch":
        return ElasticsearchSource(config)
    elif type_ == "ibm_db2":
        return IbmDb2Source(config)
    elif type_ == "athena":
        return AthenaSource(config)
    elif type_ == "synapse":
        return SynapseSource(config)

    else:
        raise ValueError(f"Unsupported source type: {type_}")
